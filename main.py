from fastapi import FastAPI, Request, HTTPException, Depends
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from telethon import TelegramClient, events
import asyncio
import os
import sys
import io
import openpyxl
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Konfigurasi Database Supabase dari Environment Variables Railway
DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    raise ValueError("DATABASE_URL belum diatur di Environment Variables Railway!")

engine = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(bind=engine)

EXCEL_REPORT_FILE = 'laporan_klaim_koin_enterprise.xlsx'

# ==========================================
# 0. SISTEM BUFFER LIVE LOG TERMINAL
# ==========================================
class LogBuffer(io.StringIO):
    def __init__(self):
        super().__init__()
        self.log_content = []

    def write(self, s):
        sys.__stdout__.write(s)
        sys.__stdout__.flush()
        if s.strip():
            formatted_s = s if s.endswith('\n') else s + '\n'
            self.log_content.append(formatted_s)
            if len(self.log_content) > 150: 
                self.log_content.pop(0)

    def flush(self):
        pass

    def get_logs(self):
        return "".join(self.log_content)

log_stream = LogBuffer()
sys.stdout = log_stream


# ==========================================
# 1. MANAJEMEN SUPABASE DB & EXCEL
# ==========================================
def inisialisasi_tabel_db():
    with SessionLocal() as db:
        db.execute(text("""
            CREATE TABLE IF NOT EXISTS users (
                username TEXT PRIMARY KEY,
                password TEXT NOT NULL,
                role TEXT DEFAULT 'user',
                status TEXT DEFAULT 'pending',
                saldo FLOAT DEFAULT 0.0
            );
        """))
        db.execute(text("""
            CREATE TABLE IF NOT EXISTS settings (
                username TEXT PRIMARY KEY,
                api_id BIGINT,
                api_hash TEXT,
                nomor_hp TEXT,
                bot_target TEXT,
                delay_aksi FLOAT DEFAULT 1.5,
                delay_repeat FLOAT DEFAULT 3.0,
                list_cookie TEXT
            );
        """))
        db.commit()

# Jalankan inisialisasi tabel saat startup
inisialisasi_tabel_db()

def ambil_dan_hapus_cookie(username):
    try:
        with SessionLocal() as db:
            res = db.execute(text("SELECT list_cookie FROM settings WHERE username = :u"), {"u": username}).fetchone()
            if not res or not res.list_cookie:
                return None
            
            cookies_str = res.list_cookie.strip()
            if not cookies_str:
                return None
                
            cookies_list = cookies_str.split('\n')
            cookie_terpakai = cookies_list.pop(0)
            sisa_cookie = '\n'.join([c.strip() for c in cookies_list if c.strip()])
            
            db.execute(text("UPDATE settings SET list_cookie = :lc WHERE username = :u"), {"lc": sisa_cookie, "u": username})
            db.commit()
            return cookie_terpakai.strip()
    except Exception as e:
        print(f"\n❌ Error ambil cookie [{username}]: {e}")
        return None

def inisialisasi_excel_jika_belum_ada():
    try:
        if not os.path.exists(EXCEL_REPORT_FILE):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Log Klaim Sukses"
            headers = ["Username", "Raw Cookie", "Jumlah Koin"]
            header_fill = openpyxl.styles.PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            header_font = openpyxl.styles.Font(bold=True, size=11)
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 15
            wb.save(EXCEL_REPORT_FILE)
    except Exception as e:
        print(f"\n❌ Gagal membuat Excel: {e}")

inisialisasi_excel_jika_belum_ada()

def catat_sukses_claim_ke_excel(username, cookie_str):
    try:
        inisialisasi_excel_jika_belum_ada()
        wb = openpyxl.load_workbook(EXCEL_REPORT_FILE)
        ws = wb.active
        
        for row in range(2, ws.max_row + 1):
            existing_cookie = ws.cell(row=row, column=2).value
            if existing_cookie and cookie_str and str(existing_cookie).strip() == str(cookie_str).strip():
                return False

        next_row = ws.max_row + 1
        border_thin = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
        
        ws.cell(row=next_row, column=1, value=username).border = border_thin
        ws.cell(row=next_row, column=2, value=cookie_str).border = border_thin
        ws.cell(row=next_row, column=3, value=3000).border = border_thin
        wb.save(EXCEL_REPORT_FILE)
        return True
    except Exception as e:
        print(f"\n❌ Gagal mencatat ke Excel: {e}")
        return False


# ==========================================
# 2. MULTI-USER TELEGRAM SESSIONS ENGINE
# ==========================================
active_clients = {}

async def jalankan_siklus_misi(username, bot_target):
    user_session = active_clients.get(username)
    if not user_session or not user_session.get("auto_claim"):
        return
        
    # Cek apakah sedang dalam proses pengiriman agar tidak nabrak ganda
    if user_session.get("lock_mission", False):
        return
    user_session["lock_mission"] = True

    try:
        config = baca_konfigurasi()
        # Ambil delay dari database user yang bersangkutan
        with SessionLocal() as db:
            st = db.execute(text("SELECT delay_repeat FROM settings WHERE username = :u"), {"u": username}).fetchone()
            delay_repeat = st.delay_repeat if st and st.delay_repeat else 3.0
            
        await asyncio.sleep(delay_repeat)
        
        if not user_session.get("auto_claim"):
            return

        with SessionLocal() as db:
            st = db.execute(text("SELECT list_cookie FROM settings WHERE username = :u"), {"u": username}).fetchone()
            sisa_cookie = st.list_cookie.strip() if st and st.list_cookie else ""
            
        if not sisa_cookie:
            print(f"\n❌ AUTO CLAIM [{username}] BERHENTI: Antrean cookie habis.")
            user_session["auto_claim"] = False
            return

        client = user_session["client"]
        await client.send_message(bot_target, '/mission')
    except Exception as e:
        print(f"\n❌ Gagal kirim /mission [{username}]: {e}")
    finally:
        # Buka kembali kunci setelah selesai
        user_session["lock_mission"] = False

def daftarkan_listener_user(username, client, bot_target):
    @client.on(events.NewMessage(chats=bot_target, incoming=True))
    async def proses_pesan_masuk(event):
        user_session = active_clients.get(username)
        if not user_session or not user_session.get("auto_claim"):
            return

        with SessionLocal() as db:
            st = db.execute(text("SELECT delay_aksi FROM settings WHERE username = :u"), {"u": username}).fetchone()
            delay_aksi = st.delay_aksi if st and st.delay_aksi else 1.5

        teks_masuk = event.raw_text
        print(f"\n📥 [{username}] Pesan masuk: {teks_masuk}")
        
        await asyncio.sleep(1)
        try:
            markup = await event.get_reply_markup()
        except:
            markup = None

        buttons = event.message.buttons or (markup.rows if markup else None)

        if buttons:
            ditemukan_misi, ditemukan_login = False, False
            for row in buttons:
                row_buttons = row.buttons if hasattr(row, 'buttons') else row
                for button in row_buttons:
                    btn_text = getattr(button, 'text', '')
                    if "Install ShopeePay" in btn_text: ditemukan_misi = True
                    elif "Raw Cookie" in btn_text: ditemukan_login = True
            
            if ditemukan_misi:
                await asyncio.sleep(delay_aksi)
                try: await event.click(0, 0)
                except: pass
            elif ditemukan_login:
                await asyncio.sleep(delay_aksi)
                try: await event.click(0, 1)
                except: pass
                    
        elif "Masukkan raw cookie" in teks_masuk:
            await asyncio.sleep(delay_aksi) 
            with SessionLocal() as db:
                st = db.execute(text("SELECT list_cookie FROM settings WHERE username = :u"), {"u": username}).fetchone()
                cookies_str_preview = st.list_cookie.strip() if st and st.list_cookie else ""
            cookie_akurat = cookies_str_preview.split('\n')[0].strip() if cookies_str_preview else ""
            
            cookie = ambil_dan_hapus_cookie(username)
            if cookie:
                await event.reply(cookie) 
                client._cookie_sedang_diproses = cookie_akurat
            else:
                user_session["auto_claim"] = False
                
        else:
            if user_session.get("auto_claim"):
                await asyncio.sleep(1.5)
                try:
                    pesan_terakhir = await client.get_messages(bot_target, limit=1)
                    teks_gabungan = pesan_terakhir[0].raw_text.replace('\n', ' | ') if pesan_terakhir else teks_masuk.replace('\n', ' | ')
                except:
                    teks_gabungan = teks_masuk.replace('\n', ' | ')

                print(f"\n🎯 [{username}] [HASIL AKHIR]: {teks_gabungan}")
                try:
                    akun_target = "Unknown"
                    if "Login:" in teks_gabungan:
                        temp = teks_gabungan.split("Login:")[1].strip()
                        akun_target = temp.split(" ")[0].split("|")[0].strip()

                    cookie_user = getattr(client, '_cookie_sedang_diproses', 'N/A')
                    berhasil_catat = catat_sukses_claim_ke_excel(akun_target if akun_target != "Unknown" else "Akun_Processed", cookie_user)
                    
                    if berhasil_catat:
                        with SessionLocal() as db:
                            usr_row = db.execute(text("SELECT saldo FROM users WHERE username = :u"), {"u": username}).fetchone()
                            if usr_row:
                                new_saldo = max(0.0, usr_row.saldo - 200.0)
                                db.execute(text("UPDATE users SET saldo = :s WHERE username = :u"), {"s": new_saldo, "u": username})
                                db.commit()
                                print(f"💰 [{username}] Saldo terpotong Rp 200 (Sisa: Rp {new_saldo})")
                        
                except Exception as parse_err:
                    print(f"⚠️ Error parsing [{username}]: {parse_err}")
                
                await jalankan_siklus_misi(username, bot_target)


# ==========================================
# 3. ENDPOINTS API REST
# ==========================================
class UserAuth(BaseModel):
    username: str
    password: str

class UserRegister(BaseModel):
    username: str
    password: str

class AdminManage(BaseModel):
    target_username: str
    status: str 
    tambah_saldo: float 

class AdminDelete(BaseModel):
    target_username: str

class ConnectTelegram(BaseModel):
    username: str
    api_id: int
    api_hash: str
    nomor_hp: str
    bot_target: str

class VerifyOTP(BaseModel):
    username: str
    kode_otp: str

class EngineSettings(BaseModel):
    username: str
    api_id: int
    api_hash: str
    nomor_hp: str
    bot_target: str
    delay_aksi: float = 1.5
    delay_repeat: float = 3.0
    list_cookie: str = ""

@app.get("/", response_class=HTMLResponse)
async def baca_index():
    if os.path.exists("index.html"):
        with open("index.html", "r", encoding="utf-8") as f:
            return f.read()
    return "<h3>File index.html belum di-upload.</h3>"

@app.post("/api/login")
async def login_user(data: UserAuth):
    with SessionLocal() as db:
        # Auto insert owner default jika tabel kosong
        owner_check = db.execute(text("SELECT username FROM users WHERE username = 'owner'")).fetchone()
        if not owner_check:
            db.execute(text("INSERT INTO users (username, password, role, status, saldo) VALUES ('owner', 'ownerpassword', 'owner', 'aktif', 100000.0) ON CONFLICT DO NOTHING"))
            db.commit()

        user = db.execute(text("SELECT * FROM users WHERE username = :u AND password = :p"), {"u": data.username, "p": data.password}).fetchone()
        if user:
            if user.role != "owner" and user.status != "aktif":
                return {"status": "error", "message": "Akun Anda belum aktif! Silakan contact Admin untuk berlangganan."}
            return {"status": "success", "role": user.role, "saldo": user.saldo, "username": data.username}
    return {"status": "error", "message": "Username atau Password salah!"}

@app.post("/api/register")
async def register_user(data: UserRegister):
    with SessionLocal() as db:
        existing = db.execute(text("SELECT username FROM users WHERE username = :u"), {"u": data.username}).fetchone()
        if existing:
            return {"status": "error", "message": "Username sudah terdaftar!"}
        
        db.execute(
            text("INSERT INTO users (username, password, role, status, saldo) VALUES (:u, :p, 'user', 'pending', 0.0)"),
            {"u": data.username, "p": data.password}
        )
        db.commit()
    return {"status": "success", "message": "Registrasi berhasil! Akun Anda berstatus PENDING. Silakan contact Admin agar akun Anda diaktifkan."}

@app.get("/api/admin/users")
async def get_all_users(username: str):
    with SessionLocal() as db:
        admin = db.execute(text("SELECT role FROM users WHERE username = :u"), {"u": username}).fetchone()
        if not admin or admin.role != "owner":
            raise HTTPException(status_code=403, detail="Akses ditolak.")
        
        rows = db.execute(text("SELECT username, role, status, saldo FROM users")).fetchall()
        users_dict = {}
        for r in rows:
            users_dict[r.username] = {
                "role": r.role,
                "status": r.status,
                "saldo": r.saldo
            }
    return {"users": users_dict}

@app.post("/api/admin/manage")
async def admin_manage_user(data: AdminManage, admin_user: str):
    with SessionLocal() as db:
        admin = db.execute(text("SELECT role FROM users WHERE username = :u"), {"u": admin_user}).fetchone()
        if not admin or admin.role != "owner":
            raise HTTPException(status_code=403, detail="Akses ditolak.")
        
        target_row = db.execute(text("SELECT saldo FROM users WHERE username = :u"), {"u": data.target_username}).fetchone()
        if not target_row:
            return {"status": "error", "message": "User tidak ditemukan."}
        
        new_saldo = target_row.saldo + data.tambah_saldo if data.tambah_saldo > 0 else target_row.saldo
        db.execute(text("UPDATE users SET status = :st, saldo = :sd WHERE username = :u"), 
                   {"st": data.status, "sd": new_saldo, "u": data.target_username})
        db.commit()
    return {"status": "success", "message": f"Data user {data.target_username} berhasil diperbarui!"}

@app.post("/api/admin/delete")
async def admin_delete_user(data: AdminDelete, admin_user: str):
    with SessionLocal() as db:
        admin = db.execute(text("SELECT role FROM users WHERE username = :u"), {"u": admin_user}).fetchone()
        if not admin or admin.role != "owner":
            raise HTTPException(status_code=403, detail="Akses ditolak.")
        
        if data.target_username == "owner":
            return {"status": "error", "message": "Akun owner utama tidak dapat dihapus!"}
        
        db.execute(text("DELETE FROM users WHERE username = :u"), {"u": data.target_username})
        db.execute(text("DELETE FROM settings WHERE username = :u"), {"u": data.target_username})
        db.commit()
    return {"status": "success", "message": f"User {data.target_username} berhasil dihapus!"}

@app.get("/api/logs")
async def get_logs():
    return {"logs": log_stream.get_logs()}

@app.get("/api/get_config")
async def get_config(username: str):
    with SessionLocal() as db:
        res = db.execute(text("SELECT * FROM settings WHERE username = :u"), {"u": username}).fetchone()
        if not res:
            return {}
        return {
            "api_id": res.api_id,
            "api_hash": res.api_hash,
            "nomor_hp": res.nomor_hp,
            "bot_target": res.bot_target,
            "delay_aksi": res.delay_aksi,
            "delay_repeat": res.delay_repeat,
            "list_cookie": res.list_cookie
        }

@app.get("/api/download_excel")
async def download_excel():
    inisialisasi_excel_jika_belum_ada()
    if os.path.exists(EXCEL_REPORT_FILE):
        return FileResponse(EXCEL_REPORT_FILE, filename='laporan_klaim_koin_enterprise.xlsx')
    return {"status": "error", "message": "File belum tersedia."}

@app.post("/api/simpan_pengaturan")
async def simpan_pengaturan(settings: EngineSettings):
    with SessionLocal() as db:
        db.execute(text("""
            INSERT INTO settings (username, api_id, api_hash, nomor_hp, bot_target, delay_aksi, delay_repeat, list_cookie)
            VALUES (:u, :ai, :ah, :nh, :bt, :da, :dr, :lc)
            ON CONFLICT (username) DO UPDATE 
            SET api_id=:ai, api_hash=:ah, nomor_hp=:nh, bot_target=:bt, delay_aksi=:da, delay_repeat=:dr, list_cookie=:lc
        """), {
            "u": settings.username,
            "ai": settings.api_id,
            "ah": settings.api_hash,
            "nh": settings.nomor_hp,
            "bt": settings.bot_target,
            "da": settings.delay_aksi,
            "dr": settings.delay_repeat,
            "lc": settings.list_cookie
        })
        db.commit()
    return {"status": "success", "message": "Konfigurasi disimpan ke Supabase!"}

@app.post("/api/hubungkan_user")
async def hubungkan_user_telegram(data: ConnectTelegram):
    try:
        session_name = f"sesi_{data.username}"
        if data.username in active_clients:
            try:
                if active_clients[data.username]["client"].is_connected():
                    await active_clients[data.username]["client"].disconnect()
            except:
                pass

        client = TelegramClient(session_name, int(data.api_id), str(data.api_hash))
        await client.connect()
        
        if not await client.is_user_authorized():
            res_otp = await client.send_code_request(data.nomor_hp)
            active_clients[data.username] = {
                "client": client,
                "phone_code_hash": res_otp.phone_code_hash,
                "auto_claim": False,
                "bot_target": data.bot_target
            }
            return {"status": "otp_required", "message": f"OTP terkirim ke Telegram [{data.username}]."}

        daftarkan_listener_user(data.username, client, data.bot_target)
        active_clients[data.username] = {
            "client": client,
            "phone_code_hash": None,
            "auto_claim": False,
            "bot_target": data.bot_target
        }
        return {"status": "connected", "message": f"Telegram [{data.username}] berhasil terhubung!"}
    except Exception as e:
        return {"status": "error", "message": f"Gagal koneksi: {str(e)}"}

@app.post("/api/verifikasi_otp_user")
async def verifikasi_otp_user(data: VerifyOTP):
    user_session = active_clients.get(data.username)
    if not user_session:
        return {"status": "error", "message": "Sesi tidak ditemukan. Hubungkan ulang."}
    
    with SessionLocal() as db:
        st = db.execute(text("SELECT nomor_hp FROM settings WHERE username = :u"), {"u": data.username}).fetchone()
        nomor_hp = st.nomor_hp if st else ""

    try:
        client = user_session["client"]
        phone_hash = user_session["phone_code_hash"]
        
        await client.sign_in(phone=nomor_hp, code=data.kode_otp.strip(), phone_code_hash=phone_hash)
        daftarkan_listener_user(data.username, client, user_session["bot_target"])
        user_session["phone_code_hash"] = None
        return {"status": "success", "message": f"Verifikasi OTP [{data.username}] Berhasil!"}
    except Exception as e:
        return {"status": "error", "message": f"Verifikasi gagal: {str(e)}"}

@app.post("/api/start_claim_user")
async def start_claim_user(data: dict):
    username = data.get("username")
    user_session = active_clients.get(username)
    if not user_session or not user_session["client"].is_connected():
        return {"status": "error", "message": "Telegram belum terhubung untuk user ini!"}
        
    with SessionLocal() as db:
        usr = db.execute(text("SELECT saldo FROM users WHERE username = :u"), {"u": username}).fetchone()
        if not usr or usr.saldo <= 0:
            return {"status": "error", "message": "Saldo Anda habis! Silakan lakukan pengisian saldo (top-up) ke Admin."}

    user_session["auto_claim"] = True
    asyncio.create_task(jalankan_siklus_misi(username, user_session["bot_target"]))
    return {"status": "success", "message": f"Automation [{username}] berhasil dijalankan."}

@app.post("/api/stop_claim_user")
async def stop_claim_user(data: dict):
    username = data.get("username")
    if username in active_clients:
        active_clients[username]["auto_claim"] = False
    return {"status": "success", "message": f"Automation [{username}] dihentikan."}
